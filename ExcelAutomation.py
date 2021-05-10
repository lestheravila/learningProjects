import openpyxl

inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']

# company=key product=value - dictionary
product_per_supplier = {}
#
total_value_per_supplier = {}

products_under_10_inv = {}
#def supplier_product():


# to check each row for supplier name
for product_row in range(2, product_list.max_row + 1):
    # assigning the the supplier column
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)


    # how many product supplier have - creating dictionary
    if supplier_name in product_per_supplier:
        # assign to new variable
        current_num_products = product_per_supplier[supplier_name]
        # increment number of product
        product_per_supplier[supplier_name] = current_num_products + 1
        # create new entry to dictionary for new supplier and set product to 1
    else:
        product_per_supplier[supplier_name] = 1

        # Calculate total inventory value per supplier
    if supplier_name in total_value_per_supplier:  # To check is it new or older supplier in dictionary
        # assign to new variable
        current_total_value = total_value_per_supplier.get(supplier_name)
        # adding
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic - products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price



print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save('inventory_with_total_value.xlsx')









